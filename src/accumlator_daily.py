"""
Axis Bank Transaction Alert Extractor
--------------------------------------
Fetches "transaction alert for Axis Bank A/c" emails from Gmail between two dates,
parses the debited/credited amount, account number, date/time, and transaction
info (particulars), then writes them to an Excel file with columns:

    Tran Date | PARTICULARS | DR | CR | BAL

SETUP (one-time):
1. Go to https://console.cloud.google.com/ -> create/select a project.
2. Enable the "Gmail API" for that project.
3. Create OAuth Client ID credentials (type: Desktop App).
4. Download the JSON and save it as `credentials.json` next to this script.
5. pip install --upgrade google-api-python-client google-auth-httplib2 google-auth-oauthlib openpyxl

USAGE:
    python axis_bank_extractor.py --start 2026-06-01 --end 2026-06-24 --out transactions.xlsx

On first run, a browser window will open asking you to log in to your Google
account and grant read-only Gmail access. A `token.json` is saved afterwards
so you won't need to log in again until it expires.
"""

import argparse
import base64
import os
import re
import sys
from datetime import datetime, date

from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from google_auth_oauthlib.flow import InstalledAppFlow
from google.auth.transport.requests import Request
from google.oauth2.credentials import Credentials

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import streamlit as st

SCOPES = ["https://www.googleapis.com/auth/gmail.readonly"]
TOKEN_FILE = "token.json"
CREDENTIALS_FILE = "credentials.json"

SENDER_QUERY = "alerts@axis.bank.in"
SUBJECT_QUERY = "transaction alert for Axis Bank"


def get_gmail_service():
    """Authenticate with Gmail API and return a service object."""
    creds = None
    if os.path.exists(TOKEN_FILE):
        creds = Credentials.from_authorized_user_file(TOKEN_FILE, SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            if not os.path.exists(CREDENTIALS_FILE):
                sys.exit(
                    f"Missing {CREDENTIALS_FILE}. Download OAuth client credentials "
                    f"from Google Cloud Console and save them with that filename."
                )
            flow = InstalledAppFlow.from_client_secrets_file(CREDENTIALS_FILE, SCOPES)
            creds = flow.run_local_server(port=0)
        with open(TOKEN_FILE, "w") as f:
            f.write(creds.to_json())
    return build("gmail", "v1", credentials=creds)


def build_gmail_query(start_date: str, end_date: str) -> str:
    """
    Gmail search query. Filters by sender and date range only — NOT subject,
    since Axis Bank uses several different subject lines for transaction
    alerts (e.g. "transaction alert for Axis Bank A/c", plain "Alert", etc.).
    Subject-agnostic filtering happens later via parse_transaction(), which
    only keeps emails whose body actually contains a debit/credit pattern.

    Gmail's `before:` is exclusive of that day, so we bump end_date by one
    day to make the range inclusive.
    start_date / end_date format: YYYY-MM-DD
    """
    from datetime import timedelta

    start_fmt = datetime.strptime(start_date, "%Y-%m-%d").strftime("%Y/%m/%d")
    end_dt = datetime.strptime(end_date, "%Y-%m-%d")
    end_fmt_exclusive = (end_dt + timedelta(days=1)).strftime("%Y/%m/%d")

    query = f"from:({SENDER_QUERY}) after:{start_fmt} before:{end_fmt_exclusive}"
    return query


def list_message_ids(service, query: str):
    """Return all Gmail message IDs matching the query, handling pagination."""
    message_ids = []
    page_token = None
    while True:
        resp = (
            service.users()
            .messages()
            .list(userId="me", q=query, pageToken=page_token, maxResults=500)
            .execute()
        )
        message_ids.extend(m["id"] for m in resp.get("messages", []))
        page_token = resp.get("nextPageToken")
        if not page_token:
            break
    return message_ids


def get_message_body_text(service, msg_id: str) -> str:
    """Fetch a message and return its plain-text (or stripped HTML) body."""
    msg = (
        service.users()
        .messages()
        .get(userId="me", id=msg_id, format="full")
        .execute()
    )
    payload = msg.get("payload", {})
    text_parts = []

    def walk_parts(part):
        mime_type = part.get("mimeType", "")
        body = part.get("body", {})
        data = body.get("data")
        if data:
            decoded = base64.urlsafe_b64decode(data).decode("utf-8", errors="ignore")
            if mime_type == "text/plain":
                text_parts.append(("plain", decoded))
            elif mime_type == "text/html":
                text_parts.append(("html", decoded))
        for sub in part.get("parts", []) or []:
            walk_parts(sub)

    walk_parts(payload)

    # Prefer plain text; fall back to stripped HTML
    plain = next((t for kind, t in text_parts if kind == "plain"), None)
    if plain:
        return plain

    html = next((t for kind, t in text_parts if kind == "html"), None)
    if html:
        return re.sub(r"<[^>]+>", " ", html)

    return ""


def parse_transaction(body: str):
    """
    Extract Tran Date, amount, DR/CR, and particulars from the email body.
    Returns a dict or None if the body doesn't match the expected pattern.
    """
    text = re.sub(r"[ \t]+", " ", body)
    text = re.sub(r"\r\n|\r", "\n", text)

    amount_match = re.search(
        r"Amount\s+(Credited|Debited)\s*:?\s*INR\s*([\d,]+\.\d{2})", text, re.IGNORECASE
    )
    if not amount_match:
        amount_match = re.search(
            r"INR\s*([\d,]+\.\d{2})\s*was\s*(credited|debited)", text, re.IGNORECASE
        )
        if amount_match:
            amount_str, direction = amount_match.group(1), amount_match.group(2)
        else:
            return None
    else:
        direction, amount_str = amount_match.group(1), amount_match.group(2)

    amount = float(amount_str.replace(",", ""))
    direction = direction.lower()

    date_match = re.search(
        r"Date\s*&\s*Time\s*:?\s*([\d]{2}-[\d]{2}-[\d]{2,4}),?\s*([\d:]{5,8})\s*IST",
        text,
        re.IGNORECASE,
    )
    tran_date = date_match.group(1) if date_match else ""

    info_match = re.search(
        r"Transaction\s+Info\s*:?\s*([^\n]+)", text, re.IGNORECASE
    )
    particulars = info_match.group(1).strip() if info_match else ""

    bal_match = re.search(
        r"(?:Available\s+Balance|Avl\s*Bal|Balance)\s*:?\s*INR?\s*([\d,]+\.\d{2})",
        text,
        re.IGNORECASE,
    )
    balance = bal_match.group(1) if bal_match else ""

    acct_match = re.search(r"A/c\s*(?:no\.?)?\s*[:\-]?\s*(XX\d+)", text, re.IGNORECASE)
    account = acct_match.group(1) if acct_match else ""

    return {
        "tran_date": tran_date,
        "particulars": particulars or account,
        "dr": amount if direction == "debited" else None,
        "cr": amount if direction == "credited" else None,
        "bal": balance,
    }


def parse_email_date(internal_date_ms: str) -> str:
    """Convert Gmail's internalDate (ms since epoch) to DD-MM-YYYY for sorting fallback."""
    dt = datetime.fromtimestamp(int(internal_date_ms) / 1000)
    return dt.strftime("%d-%m-%Y")


def fetch_axis_transactions(start_date: str, end_date: str):
    """
    Main extraction routine.
    start_date / end_date: 'YYYY-MM-DD' strings.
    Returns a list of transaction dicts.
    """
    service = get_gmail_service()
    query = build_gmail_query(start_date, end_date)
    print(f"Gmail search query: {query}")

    msg_ids = list_message_ids(service, query)
    print(f"Found {len(msg_ids)} matching emails.")

    transactions = []
    for mid in msg_ids:
        try:
            msg = service.users().messages().get(userId="me", id=mid, format="full").execute()
            body = get_message_body_text(service, mid)
            parsed = parse_transaction(body)
            if parsed is None:
                continue
            if not parsed["tran_date"]:
                parsed["tran_date"] = parse_email_date(msg.get("internalDate", "0"))
            transactions.append(parsed)
        except HttpError as e:
            print(f"Skipping message {mid} due to API error: {e}")

    def sort_key(t):
        for fmt in ("%d-%m-%y", "%d-%m-%Y"):
            try:
                return datetime.strptime(t["tran_date"], fmt)
            except ValueError:
                continue
        return datetime.min

    transactions.sort(key=sort_key)
    return transactions


def write_to_excel(transactions, out_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Axis Transactions"

    headers = ["Tran Date", "PARTICULARS", "DR", "CR", "BAL"]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for t in transactions:
        ws.append([
            t["tran_date"],
            t["particulars"],
            t["dr"] if t["dr"] is not None else None,
            t["cr"] if t["cr"] is not None else None,
            t["bal"] if t["bal"] else None,
        ])

    for row in ws.iter_rows(min_row=2, min_col=3, max_col=5):
        for cell in row:
            if cell.value is not None:
                cell.number_format = "#,##0.00"

    widths = [14, 50, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save(out_path)
    print(f"Saved {len(transactions)} transactions to {out_path}")



if __name__ == "__main__":
    transactions = fetch_axis_transactions('2026-06-13', date.today().strftime("%Y-%m-%d"))
    st.dataframe(transactions)